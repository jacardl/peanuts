import multiprocessing as mp
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

import api
from common import *
import var as v


class MemMonitor(threading.Thread):
    def __init__(self, interval=5, period=1):
        threading.Thread.__init__(self)
        self.interval = interval
        self.period = period
        self.running = False
        self.callback = None
        self.terminal = ShellCommand(v.CONNECTION_TYPE)
        self.terminal.connect(v.HOST, v.USR, v.PASSWD)
        self.tmpHost = v.HOST

    def run(self):
        self.running = True
        self.plot = []
        last_time = t.time()
        while self.running:
            curr_time = t.time()
            if curr_time - last_time >= self.interval:
                last_time = curr_time
                try:
                    if v.HOST != self.tmpHost:
                        self.terminal.connect(v.HOST, v.USR, v.PASSWD)
                        self.tmpHost = v.HOST
                    res = self.terminal.getTotalMemInfo()
                except Exception, e:
                    continue
                memUsed = res["used"]
                self.plot.append(memUsed)
                if self.callback is not None:
                    self.callback.after_read_res(res, t)
            t.sleep(self.period)
        self.terminal.close()
        p = mp.Process(target=drawMem, args=(self.plot,))
        p.start()
        p.join()

    def stop(self):
        self.running = False


class HttpMemMonitor(threading.Thread):
    def __init__(self, interval=5, period=1):
        threading.Thread.__init__(self)
        self.interval = interval
        self.period = period
        self.running = False
        self.callback = None
        self.terminal = api.HttpClient()
        self.terminal.connect(host=v.HOST, password=v.WEB_PWD)
        self.tmpHost = v.HOST

    def run(self):
        self.running = True
        self.plot = []
        last_time = t.time()
        while self.running:
            curr_time = t.time()
            if curr_time - last_time >= self.interval:
                last_time = curr_time
                try:
                    if v.HOST != self.tmpHost:
                        self.terminal.connect(host=v.HOST, password=v.WEB_PWD)
                        self.tmpHost = v.HOST
                    memUsed = api.getDeviceMem(self.terminal, v.DEVICE_STATUS_LOG)
                except Exception, e:
                    continue
                if memUsed is not None:
                    self.plot.append(memUsed)
                if self.callback is not None:
                    self.callback.after_read_res(memUsed, t)
            t.sleep(self.period)
        self.terminal.close()
        p = mp.Process(target=drawMem, args=(self.plot,))
        p.start()
        p.join()

    def stop(self):
        self.running = False


class HttpCPUMonitor(threading.Thread):
    def __init__(self, interval=5, period=1):
        threading.Thread.__init__(self)
        self.interval = interval
        self.period = period
        self.running = False
        self.callback = None
        self.terminal = api.HttpClient()
        self.terminal.connect(host=v.HOST, password=v.WEB_PWD)
        self.tmpHost = v.HOST

    def run(self):
        self.running = True
        self.plot = []
        last_time = t.time()
        while self.running:
            curr_time = t.time()
            if curr_time - last_time >= self.interval:
                last_time = curr_time
                try:
                    if v.HOST != self.tmpHost:
                        self.terminal.connect(host=v.HOST, password=v.WEB_PWD)
                        self.tmpHost = v.HOST
                    cpuLoad = api.getDeviceCPU(self.terminal, v.DEVICE_STATUS_LOG)
                except Exception, e:
                    continue
                if cpuLoad is not None:
                    self.plot.append(cpuLoad)
                if self.callback is not None:
                    self.callback.after_read_res(cpuLoad, t)
            t.sleep(self.period)
        self.terminal.close()
        p = mp.Process(target=drawCPU, args=(self.plot,))
        p.start()
        p.join()

    def stop(self):
        self.running = False


class HttpMemCPUMonitor(threading.Thread):
    def __init__(self, interval=5, period=1):
        threading.Thread.__init__(self)
        self.interval = interval
        self.period = period
        self.running = False
        self.callback = None
        self.terminal = api.HttpClient()
        self.terminal.connect(host=v.HOST, password=v.WEB_PWD)
        self.tmpHost = v.HOST

    def run(self):
        self.running = True
        self.plot = []
        self.plot2 = []
        status = dict()
        last_time = t.time()
        while self.running:
            curr_time = t.time()
            if curr_time - last_time >= self.interval:
                last_time = curr_time
                try:
                    if v.HOST != self.tmpHost:
                        self.terminal.connect(host=v.HOST, password=v.WEB_PWD)
                        self.tmpHost = v.HOST
                    status = api.getDeviceSystemInfo(self.terminal, v.DEVICE_STATUS_LOG)
                except Exception, e:
                    continue
                if status.get('memUsed') is not None:
                    self.plot.append(status.get('memUsed'))
                if status.get('cpuLoad') is not None:
                    self.plot2.append(status.get('cpuLoad'))
                if self.callback is not None:
                    pass
            t.sleep(self.period)
        self.terminal.close()
        p = mp.Process(target=drawMem, args=(self.plot,))
        p2 = mp.Process(target=drawCPU, args=(self.plot2,))
        p.start()
        p2.start()
        p.join()
        p2.join()

    def stop(self):
        self.running = False


class MemMonitorXlsx(threading.Thread):
    def __init__(self, interval, count=0, file="temp.xlsx"):
        threading.Thread.__init__(self)
        self.interval = interval
        self.count = count
        self.running = False
        self.callback = None
        self.file = file
        self.sheetDaemon = "User Mem Tracking"
        self.sheetKernel = "Kernel Mem Tracking"
        self.terminal = ShellCommand(v.CONNECTION_TYPE)
        self.ret = self.terminal.connect(v.HOST, v.USR, v.PASSWD)

    def run(self):
        self.running = True
        # daemon = threading.Thread(target=daemonKernelMonitor, args=(self.terminal, self.interval, self.count,self.file, self.sheetDaemon, self.sheetKernel))
        # def daemonKernelMonitor(terminal, interval, count, filename, sheet1, sheet2):
        # daemon.setDaemon(True)
        # daemon.start()
        # while self.running and daemon.isAlive():
        #     daemon.join(1)

        tmpHost = v.HOST
        last_time = t.time()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = self.sheetDaemon
        ws2 = wb.create_sheet(title=self.sheetKernel)
        wb.save(self.file)
        keys1 = list()
        keys2 = list()
        daemonNameCol = 3  # col 1 write time, col 2 write total mem, so vsz begin with 3
        kernelCacheNameCol = 3
        vszRow = 2  # Row 1 write title, value begin with Row 2
        ws1.cell(row=1, column=2).value = "Total Used"  # write Total used title
        ws2.cell(row=1, column=2).value = "Total Cache"  # write Total cache title
        keys2.append("Total Cache")
        if self.count >= 1:
            curr_count = 1
            while curr_count <= self.count:
                curr_time = t.time()
                if curr_time - last_time >= self.interval:
                    last_time = curr_time
                    try:
                        if v.HOST != tmpHost:
                            self.terminal.connect(v.HOST, v.USR, v.PASSWD)
                            tmpHost = v.HOST
                        totalmem = self.terminal.getTotalMemInfo()
                        time = self.terminal.getTimeStr()
                        newDic1 = getDaemonRss(self.terminal)
                        newDic2 = getKernelSlab(self.terminal)
                    except Exception, e:
                        curr_count += 1
                        continue
                    newKeys1 = newDic1.keys()
                    newKeys2 = newDic2.keys()

                    ws1.cell(row=vszRow, column=1).value = time  # write time
                    ws1.cell(row=vszRow, column=2).value = totalmem["used"]  # write total mem
                    ws2.cell(row=vszRow, column=1).value = time
                    for i in newKeys1:
                        try:
                            vszCol = keys1.index(i) + 3  # begin from col 3
                            ws1.cell(row=vszRow, column=vszCol).value = newDic1[i]  # write daemon vsz

                        except:
                            flag = 0
                            for e in v.EXCEPTIONS:  # exclude some daemons
                                if e not in i:
                                    pass
                                else:
                                    flag = 1
                                    break
                            if flag == 0:
                                ws1.cell(row=1, column=daemonNameCol).value = i  # write new daemon title
                                ws1.cell(row=vszRow, column=daemonNameCol).value = newDic1[i]  # write new daemon vsz
                                keys1.append(i)
                                daemonNameCol += 1

                    for j in newKeys2:
                        try:
                            vszCol = keys2.index(j) + 2  # begin from col 3
                            ws2.cell(row=vszRow, column=vszCol).value = newDic2[j]  # write kernel cache size

                        except:
                            flag = 0
                            for e in v.KERNEL_EXCEPTIONS:  # exclude some daemons
                                if e not in j:
                                    pass
                                else:
                                    flag = 1
                                    break
                            if flag == 0:
                                ws2.cell(row=1, column=kernelCacheNameCol).value = j  # write new cache name
                                ws2.cell(row=vszRow, column=kernelCacheNameCol).value = newDic2[j]  # write new cache size
                                keys2.append(j)
                                kernelCacheNameCol += 1

                    vszRow += 1
                    curr_count += 1
                    try:
                        wb.save(self.file)
                    except:
                        self.terminal.close()
                        break
                t.sleep(1)
        elif self.count is 0:
            while self.running:
                curr_time = t.time()
                if curr_time - last_time >= self.interval:
                    last_time = curr_time
                    try:
                        if v.HOST != tmpHost:
                            self.terminal.connect(v.HOST, v.USR, v.PASSWD)
                            tmpHost = v.HOST
                        totalmem = self.terminal.getTotalMemInfo()
                        time = self.terminal.getTimeStr()
                        newDic1 = getDaemonRss(self.terminal)
                        newDic2 = getKernelSlab(self.terminal)
                    except Exception, e:
                        continue
                    newKeys1 = newDic1.keys()
                    newKeys2 = newDic2.keys()

                    ws1.cell(row=vszRow, column=1).value = time  # write time
                    ws1.cell(row=vszRow, column=2).value = totalmem["used"]  # write total mem
                    ws2.cell(row=vszRow, column=1).value = time
                    for i in newKeys1:
                        try:
                            vszCol = keys1.index(i) + 3  # begin from col 3
                            ws1.cell(row=vszRow, column=vszCol).value = newDic1[i]  # write daemon vsz

                        except:
                            flag = 0
                            for e in v.EXCEPTIONS:  # exclude some daemons
                                if e not in i:
                                    pass
                                else:
                                    flag = 1
                                    break
                            if flag == 0:
                                ws1.cell(row=1, column=daemonNameCol).value = i  # write new daemon title
                                ws1.cell(row=vszRow, column=daemonNameCol).value = newDic1[i]  # write new daemon vsz
                                keys1.append(i)
                                daemonNameCol += 1

                    for j in newKeys2:
                        try:
                            vszCol = keys2.index(j) + 2  # begin from col 3
                            ws2.cell(row=vszRow, column=vszCol).value = newDic2[j]  # write kernel cache size

                        except:
                            flag = 0
                            for e in v.KERNEL_EXCEPTIONS:  # exclude some daemons
                                if e not in j:
                                    pass
                                else:
                                    flag = 1
                                    break
                            if flag == 0:
                                ws2.cell(row=1, column=kernelCacheNameCol).value = j  # write new cache name
                                ws2.cell(row=vszRow, column=kernelCacheNameCol).value = newDic2[j]  # write new cache size
                                keys2.append(j)
                                kernelCacheNameCol += 1

                    vszRow += 1
                    try:
                        wb.save(self.file)
                    except:
                        self.terminal.close()
                        break
                t.sleep(1)
        memDiffCalc(self.file, [self.sheetDaemon, self.sheetKernel])
        self.stop()

    def stop(self):
        self.running = False
        self.terminal.close()


def drawCPU(data):
    # draw a chart
    fig, ax = plt.subplots(figsize=(12, 6))
    print "draw CPU chart"
    ax.plot(xrange(len(data)), data)
    # ax.set_title('Total Memory Used')
    plt.suptitle(v.MAIL_PIC4.split(".")[0].split('\\')[-1].replace('_', ' '), fontsize=12, style='oblique',
                 va='top')
    # plt.xlabel('Counts')
    plt.ylabel('Percent%')
    # plt.show()
    plt.savefig(v.MAIL_PIC4)
    plt.close()


def drawMem(data):
    # draw a chart
    fig, ax = plt.subplots(figsize=(12, 6))
    print "draw memory chart"
    ax.plot(xrange(len(data)), data)
    # ax.set_title('Total Memory Used')
    plt.suptitle("Total Memory Used")
    plt.suptitle(v.MAIL_PIC1.split(".")[0].split('\\')[-1].replace('_', ' '), fontsize=12, style='oblique',
                 va='top')
    # plt.xlabel('Counts')
    plt.ylabel('KB')
    # plt.show()
    plt.savefig(v.MAIL_PIC1)
    plt.close()


def daemonMonitor(terminal, interval, count, filename, sheetname):
    curr_count = 1
    last_time = t.time()
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    keys = []
    row1Col = 3  # col 1 write time, col 2 write total mem, so vsz begin with 3
    vszRow = 2  # Row 1 write title, value begin with Row 2
    ws.cell(row=1, column=2).value = "Total Used"  # write Total used title

    if count >= 1:
        while curr_count <= count:
            curr_time = t.time()
            if curr_time - last_time >= interval:
                last_time = curr_time
                try:
                    # terminal.getPidNameVSZDic()
                    totalmem = terminal.getTotalMemInfo()
                    time = terminal.getTimeStr()
                    newDic = getDaemonRss(terminal)
                except Exception, e:
                    curr_count += 1
                    continue
                newKeys = newDic.keys()

                ws.cell(row=vszRow, column=1).value = time  # write time
                ws.cell(row=vszRow, column=2).value = totalmem["used"]  # write total mem
                for i in newKeys:
                    try:
                        vszCol = keys.index(i) + 3  # begin from col 3
                        ws.cell(row=vszRow, column=vszCol).value = newDic[i]  # write daemon vsz

                    except:
                        flag = 0
                        for e in v.EXCEPTIONS:  # exclude some daemons
                            if e not in i:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws.cell(row=1, column=row1Col).value = i  # write new daemon title
                            # self.ws.col(row0Col).width = v.WIDTH
                            ws.cell(row=vszRow, column=row1Col).value = newDic[i]  # write new daemon vsz
                            keys.append(i)
                            row1Col += 1
                vszRow += 1
                curr_count += 1
                try:
                    wb.save(filename)
                except:
                    terminal.close()
                    break
            t.sleep(1)
    elif count is 0:
        while True:
            curr_time = t.time()
            if curr_time - last_time >= interval:
                last_time = curr_time
                try:
                    totalmem = terminal.getTotalMemInfo()
                    time = terminal.getTimeStr()
                    newDic = getDaemonRss(terminal)
                except Exception, e:
                    curr_count += 1
                    continue
                newKeys = newDic.keys()

                ws.cell(row=vszRow, column=1).value = time  # write time
                ws.cell(row=vszRow, column=2).value = totalmem["used"]  # write total mem
                for i in newKeys:
                    try:
                        vszCol = keys.index(i) + 3  # begin from col 3
                        ws.cell(row=vszRow, column=vszCol).value = newDic[i]  # write daemon vsz

                    except:
                        flag = 0
                        for e in v.EXCEPTIONS:  # exclude some daemons
                            if e not in i:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws.cell(row=1, column=row1Col).value = i  # write new daemon title
                            # self.ws.col(row0Col).width = v.WIDTH
                            ws.cell(row=vszRow, column=row1Col).value = newDic[i]  # write new daemon vsz
                            keys.append(i)
                            row1Col += 1
                vszRow += 1
                curr_count += 1
                try:
                    wb.save(filename)
                except:
                    terminal.close()
                    break
            t.sleep(1)


def kernelMonitor(terminal, interval, count, filename, sheetname):
    curr_count = 1
    last_time = t.time()
    wb = Workbook()
    ws = wb.create_sheet(title=sheetname)
    keys = [] # repersent cache name has been written
    row1Col = 3  # col 1 write time, col 2 write total cache
    vszRow = 2  # Row 1 write title, value begin with Row 2
    ws.cell(row=1, column=2).value = "Total Cache"  # write Total cache title
    keys.append("Total Cache")

    if count >= 1:
        while curr_count <= count:
            curr_time = t.time()
            if curr_time - last_time >= interval:
                last_time = curr_time
                try:
                    time = terminal.getTimeStr()
                    newDic = getKernelSlab(terminal)
                except Exception, e:
                    curr_count += 1
                    continue
                newKeys = newDic.keys()

                ws.cell(row=vszRow, column=1).value = time  # write time

                for i in newKeys:
                    try:
                        vszCol = keys.index(i) + 2  # begin from col 2
                        ws.cell(row=vszRow, column=vszCol).value = newDic[i]  # write kernel cache size

                    except:
                        flag = 0
                        for e in v.KERNEL_EXCEPTIONS:  # exclude some kernel cache name
                            if e not in i:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws.cell(row=1, column=row1Col).value = i  # write new cache name
                            ws.cell(row=vszRow, column=row1Col).value = newDic[i]  # write new cache size
                            keys.append(i)
                            row1Col += 1
                vszRow += 1
                curr_count += 1
                try:
                    wb.save(filename)
                except:
                    terminal.close()
                    break
            t.sleep(1)
    elif count is 0:
        while True:
            curr_time = t.time()
            if curr_time - last_time >= interval:
                last_time = curr_time
                try:
                    time = terminal.getTimeStr()
                    newDic = getKernelSlab(terminal)
                except Exception, e:
                    continue
                newKeys = newDic.keys()

                ws.cell(row=vszRow, column=1).value = time  # write time
                for i in newKeys:
                    try:
                        vszCol = keys.index(i) + 2  # begin from col 3
                        ws.cell(row=vszRow, column=vszCol).value = newDic[i]  # write daemon vsz

                    except:
                        flag = 0
                        for e in v.EXCEPTIONS:  # exclude some daemons
                            if e not in i:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws.cell(row=1, column=row1Col).value = i  # write new cache name
                            ws.cell(row=vszRow, column=row1Col).value = newDic[i]  # write new cache size
                            keys.append(i)
                            row1Col += 1
                vszRow += 1
                try:
                    wb.save(filename)
                except:
                    terminal.close()
                    break
            t.sleep(1)


def daemonKernelMonitor(terminal, interval, count, filename, sheet1, sheet2):
    tmpHost = v.HOST
    last_time = t.time()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet1
    ws2 = wb.create_sheet(title=sheet2)
    wb.save(filename)
    keys1 = list()
    keys2 = list()
    daemonNameCol = 3  # col 1 write time, col 2 write total mem, so vsz begin with 3
    kernelCacheNameCol = 3
    vszRow = 2  # Row 1 write title, value begin with Row 2
    ws1.cell(row=1, column=2).value = "Total Used"  # write Total used title
    ws2.cell(row=1, column=2).value = "Total Cache"  # write Total cache title
    keys2.append("Total Cache")
    if count >= 1:
        curr_count = 1
        while curr_count <= count:
            curr_time = t.time()
            if curr_time - last_time >= interval:
                last_time = curr_time
                try:
                    if v.HOST != tmpHost:
                        terminal.connect(v.HOST, v.USR, v.PASSWD)
                        tmpHost = v.HOST
                    totalmem = terminal.getTotalMemInfo()
                    time = terminal.getTimeStr()
                    newDic1 = getDaemonRss(terminal)
                    newDic2 = getKernelSlab(terminal)
                except Exception, e:
                    curr_count += 1
                    continue
                newKeys1 = newDic1.keys()
                newKeys2 = newDic2.keys()

                ws1.cell(row=vszRow, column=1).value = time  # write time
                ws1.cell(row=vszRow, column=2).value = totalmem["used"]  # write total mem
                ws2.cell(row=vszRow, column=1).value = time
                for i in newKeys1:
                    try:
                        vszCol = keys1.index(i) + 3  # begin from col 3
                        ws1.cell(row=vszRow, column=vszCol).value = newDic1[i]  # write daemon vsz

                    except:
                        flag = 0
                        for e in v.EXCEPTIONS:  # exclude some daemons
                            if e not in i:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws1.cell(row=1, column=daemonNameCol).value = i  # write new daemon title
                            ws1.cell(row=vszRow, column=daemonNameCol).value = newDic1[i]  # write new daemon vsz
                            keys1.append(i)
                            daemonNameCol += 1

                for j in newKeys2:
                    try:
                        vszCol = keys2.index(j) + 2  # begin from col 3
                        ws2.cell(row=vszRow, column=vszCol).value = newDic2[j]  # write kernel cache size

                    except:
                        flag = 0
                        for e in v.KERNEL_EXCEPTIONS:  # exclude some daemons
                            if e not in j:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws2.cell(row=1, column=kernelCacheNameCol).value = j  # write new cache name
                            ws2.cell(row=vszRow, column=kernelCacheNameCol).value = newDic2[j]  # write new cache size
                            keys2.append(j)
                            kernelCacheNameCol += 1

                vszRow += 1
                curr_count += 1
                try:
                    wb.save(filename)
                except:
                    terminal.close()
                    break
            t.sleep(1)
    elif count is 0:
        while True:
            curr_time = t.time()
            if curr_time - last_time >= interval:
                last_time = curr_time
                try:
                    if v.HOST != tmpHost:
                        terminal.connect(v.HOST, v.USR, v.PASSWD)
                        tmpHost = v.HOST
                    totalmem = terminal.getTotalMemInfo()
                    time = terminal.getTimeStr()
                    newDic1 = getDaemonRss(terminal)
                    newDic2 = getKernelSlab(terminal)
                except Exception, e:
                    continue
                newKeys1 = newDic1.keys()
                newKeys2 = newDic2.keys()

                ws1.cell(row=vszRow, column=1).value = time  # write time
                ws1.cell(row=vszRow, column=2).value = totalmem["used"]  # write total mem
                ws2.cell(row=vszRow, column=1).value = time
                for i in newKeys1:
                    try:
                        vszCol = keys1.index(i) + 3  # begin from col 3
                        ws1.cell(row=vszRow, column=vszCol).value = newDic1[i]  # write daemon vsz

                    except:
                        flag = 0
                        for e in v.EXCEPTIONS:  # exclude some daemons
                            if e not in i:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws1.cell(row=1, column=daemonNameCol).value = i  # write new daemon title
                            ws1.cell(row=vszRow, column=daemonNameCol).value = newDic1[i]  # write new daemon vsz
                            keys1.append(i)
                            daemonNameCol += 1

                for j in newKeys2:
                    try:
                        vszCol = keys2.index(j) + 2  # begin from col 3
                        ws2.cell(row=vszRow, column=vszCol).value = newDic2[j]  # write kernel cache size

                    except:
                        flag = 0
                        for e in v.KERNEL_EXCEPTIONS:  # exclude some daemons
                            if e not in j:
                                pass
                            else:
                                flag = 1
                                break
                        if flag == 0:
                            ws2.cell(row=1, column=kernelCacheNameCol).value = j  # write new cache name
                            ws2.cell(row=vszRow, column=kernelCacheNameCol).value = newDic2[j]  # write new cache size
                            keys2.append(j)
                            kernelCacheNameCol += 1

                vszRow += 1
                try:
                    wb.save(filename)
                except:
                    terminal.close()
                    break
            t.sleep(1)


def memDiffCalc(filename, sheetname):
    try:
        wb = load_workbook(filename)
    except:
        print "specified file no exists!"
        return
    # ws = wb.active
    for sheet in sheetname:
        ws = wb[sheet]
        maxRow = ws.max_row
        maxCol = ws.max_column

        for col in range(2, maxCol + 1):
            rowStart = None
            rowEnd = None
            for row in range(2, maxRow+1):
                if rowStart is None:
                    rowStart = ws.cell(row=row, column=col).value
                else:
                    rowStart = float(rowStart)
                    break
            for row in reversed(range(2, maxRow+1)):
                if rowEnd is None:
                    rowEnd = ws.cell(row=row, column=col).value
                else:
                    rowEnd = float(rowEnd)
                    break
            try:
                diff = (rowEnd - rowStart) / rowStart
            except Exception, e:
                print e, "rowStart=%s, rowEnd=%s"%(rowStart, rowEnd)
                diff = 0

            if diff != 0:
                # self.ws.write(maxRow, col, diff, stylePercent)
                sum = ws.cell(row=maxRow + 1, column=col)
                diff = "{:.1%}".format(diff)
                sum.value = diff
        wb.save(filename)





if __name__ == '__main__':

    v.CONNECTION_TYPE = 2
    v.HOST = "192.168.31.1"
    v.USR = "root"
    v.PASSWD = "admin"
    # v.HOST = '192.168.31.1'
    # v.WEB_PWD = '12345678'
    # webclient = api.HttpClient()
    # webclient.connect(host=v.HOST, password=v.WEB_PWD)
    # memMon = MemMonitorXlsx(2, count=5, file='a.xlsx')
    # memMon.start()
    # t.sleep(20)
    # memMon.stop()
    memDiffCalc('memory_tracking.xlsx', ['User Mem Tracking', 'Kernel Mem Tracking'])
