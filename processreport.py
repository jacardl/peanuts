# -*- coding: utf8 -*-
import multiprocessing as mp
import threading
import re
from collections import *
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.cell import column_index_from_string
import shutil

from var import *
import data

class ProcessReport(mp.Process):
    def __init__(self, report, q):
        mp.Process.__init__(self)
        self.running = False
        self.report = report
        self.result = OrderedDict()
        self.qu = q

    def run(self):
        self.running = True
        time = GetTimeUsed(self.report)
        test = GetTestResult(self.report)
        throughput = GetThroughputLog(self.report)
        online = GetOnlineLog(self.report)
        module = GetTestModule(self.report)
        time.start()
        test.start()
        throughput.start()
        online.start()
        module.start()
        while time.is_alive() or test.is_alive() or throughput.is_alive() \
                or online.is_alive() or module.is_alive():
            pass
        self.result.update(error=test.result['error'])
        self.result.update(ranpass=test.result["ranpass"])
        self.result.update(ransum=test.result["ransum"])
        if self.result['ransum'] == 0:
            self.result.update(percent=0)
        else:
            self.result.update(percent=self.result["ranpass"]/float(self.result['ransum'])*100)
        self.result.update(time=time.timeUsed)
        self.result.update(onlinesum=online.result["pass"]+online.result["fail"])
        self.result.update(onlinepass=online.result["pass"])
        onlineSum = online.result["pass"]+online.result["fail"]
        if onlineSum == 0:
            self.result.update(onlinepercent=0)
        else:
            self.result.update(onlinepercent=online.result["pass"]/float(onlineSum)*100)
        self.result.update(module=module.result)
        if len(throughput.wanBWResult) is not 0:
            self.result.update(wandownload=throughput.wanBWResult['wandownload'])
            self.result.update(wanupload=throughput.wanBWResult['wanupload'])
        self.qu.put(self.result)
        shutil.make_archive(self.report, 'zip', TEST_SUITE_LOG_PATH)
        self.stop()

    def stop(self):
        self.running = False


class GetTimeUsed(threading.Thread):
    def __init__(self, report):
        threading.Thread.__init__(self)
        self.running = False
        self.reportName = report
        self.timeUsed = 0

    def run(self):
        self.running = True
        f = open(self.reportName)
        for line in f:
            if not line.isspace():
                m = re.search('Ran\s+\d+\s+[tes]+\s+in\s+(\d+)', line)
                if m:
                    self.timeUsed += int(m.group(1))
        f.close()
        self.timeUsed = float(self.timeUsed/3600.0)
        self.timeUsed = round(self.timeUsed, 2)
        self.stop()

    def stop(self):
        self.running = False


class GetTestResult(threading.Thread):
    def __init__(self, report):
        threading.Thread.__init__(self)
        self.running = False
        self.reportName = report
        self.result = {
            'error': 0,
            'ranfail': 0,
            'ranpass': 0,
            'ransum': 0,
            }

    def run(self):
        self.running = True
        f = open(self.reportName)
        for line in f:
            if not line.isspace():
                m = re.search('Ran\s+(\d+)\s+[tes]+\s+in.*', line)
                if m:
                    if self.result['ransum'] is 0:
                        self.result['ransum'] = int(m.group(1))
                    next(f)
                    line2 = next(f)
                    m2 = re.search('(FAILED \((failures=(\d+))?(, )?(errors=(\d+))?\))?(OK)?', line2)
                    if m2:
                        if m.group(1) and m2.group(3) and m2.group(6):
                            self.result["ranpass"] += (int(m.group(1)) - int(m2.group(3)))
                            self.result['ranfail'] = int(m2.group(3))
                            self.result['error'] += int(m2.group(6))
                        elif m.group(1) and m2.group(3):
                            self.result["ranpass"] += (int(m.group(1)) - int(m2.group(3)))
                            self.result['ranfail'] = int(m2.group(3))
                        elif m.group(1) and m2.group(6):
                            self.result["ranpass"] += int(m.group(1))
                            self.result['error'] += int(m2.group(6))
                            self.result['ranfail'] = 0
                        elif m.group(1) and m2.group(7):
                            self.result['ranpass'] += int(m.group(1))
                            self.result['ranfail'] = 0

        f.close()
        self.stop()

    def stop(self):
        self.running = False


class GetThroughputLog(threading.Thread):
    def __init__(self, report):
        threading.Thread.__init__(self)
        self.reportName = report
        self.logPath = TEST_SUITE_LOG_PATH
        self.wanBWResult = dict()
        self.dut2g = {
            '20tx': [],
            '20rx': [],
            '40tx': [],
            '40rx': [],
        }
        self.dut5g = {
            '20tx': [],
            '40tx': [],
            '80tx': [],
            '20rx': [],
            '40rx': [],
            '80rx': [],
        }
        self.lan2g = {
            '20tx': [],
            '20rx': [],
            '40tx': [],
            '40rx': [],
        }
        self.lan5g = {
            '20tx': [],
            '40tx': [],
            '80tx': [],
            '20rx': [],
            '40rx': [],
            '80rx': [],
        }
        self.wan2g = {
            'tx':[],
            'rx':[],
            'relaytx':[],
            'relayrx':[],
            'wrelaytx':[],
            'wrelayrx':[],
            'guesttx':[],
            'guestrx':[],
        }
        self.wan5g = {
            'tx':[],
            'rx':[],
            'relaytx':[],
            'relayrx':[],
            'wrelaytx':[],
            'wrelayrx':[],
        }
        self.dut = {
            'CH1 20M': "self.dut2g['20tx'].append(speedDict['tx']), self.dut2g['20rx'].append(speedDict['rx'])",
            'CH6 20M': "self.dut2g['20tx'].append(speedDict['tx']), self.dut2g['20rx'].append(speedDict['rx'])",
            'CH11 20M': "self.dut2g['20tx'].append(speedDict['tx']), self.dut2g['20rx'].append(speedDict['rx'])",
            'CH13 20M': "self.dut2g['20tx'].append(speedDict['tx']), self.dut2g['20rx'].append(speedDict['rx'])",
            'CH1 40M': "self.dut2g['40tx'].append(speedDict['tx']), self.dut2g['40rx'].append(speedDict['rx'])",
            'CH6 40M': "self.dut2g['40tx'].append(speedDict['tx']), self.dut2g['40rx'].append(speedDict['rx'])",
            'CH11 40M': "self.dut2g['40tx'].append(speedDict['tx']), self.dut2g['40rx'].append(speedDict['rx'])",
            'CH13 40M': "self.dut2g['40tx'].append(speedDict['tx']), self.dut2g['40rx'].append(speedDict['rx'])",
            'CH36 20M': "self.dut5g['20tx'].append(speedDict['tx']), self.dut5g['20rx'].append(speedDict['rx'])",
            'CH52 20M': "self.dut5g['20tx'].append(speedDict['tx']), self.dut5g['20rx'].append(speedDict['rx'])",
            'CH149 20M': "self.dut5g['20tx'].append(speedDict['tx']), self.dut5g['20rx'].append(speedDict['rx'])",
            'CH165 20M': "self.dut5g['20tx'].append(speedDict['tx']), self.dut5g['20rx'].append(speedDict['rx'])",
            'CH36 40M': "self.dut5g['40tx'].append(speedDict['tx']), self.dut5g['40rx'].append(speedDict['rx'])",
            'CH44 40M': "self.dut5g['40tx'].append(speedDict['tx']), self.dut5g['40rx'].append(speedDict['rx'])",
            'CH52 40M': "self.dut5g['40tx'].append(speedDict['tx']), self.dut5g['40rx'].append(speedDict['rx'])",
            'CH60 40M': "self.dut5g['40tx'].append(speedDict['tx']), self.dut5g['40rx'].append(speedDict['rx'])",
            'CH149 40M': "self.dut5g['40tx'].append(speedDict['tx']), self.dut5g['40rx'].append(speedDict['rx'])",
            'CH157 40M': "self.dut5g['40tx'].append(speedDict['tx']), self.dut5g['40rx'].append(speedDict['rx'])",
            'CH36 80M': "self.dut5g['80tx'].append(speedDict['tx']), self.dut5g['80rx'].append(speedDict['rx'])",
            'CH52 80M': "self.dut5g['80tx'].append(speedDict['tx']), self.dut5g['80rx'].append(speedDict['rx'])",
            'CH149 80M': "self.dut5g['80tx'].append(speedDict['tx']), self.dut5g['80rx'].append(speedDict['rx'])",
        }
        self.lan = {
            'CH1 20M': "self.lan2g['20tx'].append(speedDict['tx']), self.lan2g['20rx'].append(speedDict['rx'])",
            'CH6 20M': "self.lan2g['20tx'].append(speedDict['tx']), self.lan2g['20rx'].append(speedDict['rx'])",
            'CH11 20M': "self.lan2g['20tx'].append(speedDict['tx']), self.lan2g['20rx'].append(speedDict['rx'])",
            'CH13 20M': "self.lan2g['20tx'].append(speedDict['tx']), self.lan2g['20rx'].append(speedDict['rx'])",
            'CH1 40M': "self.lan2g['40tx'].append(speedDict['tx']), self.lan2g['40rx'].append(speedDict['rx'])",
            'CH6 40M': "self.lan2g['40tx'].append(speedDict['tx']), self.lan2g['40rx'].append(speedDict['rx'])",
            'CH11 40M': "self.lan2g['40tx'].append(speedDict['tx']), self.lan2g['40rx'].append(speedDict['rx'])",
            'CH13 40M': "self.lan2g['40tx'].append(speedDict['tx']), self.lan2g['40rx'].append(speedDict['rx'])",
            'CH36 20M': "self.lan5g['20tx'].append(speedDict['tx']), self.lan5g['20rx'].append(speedDict['rx'])",
            'CH52 20M': "self.lan5g['20tx'].append(speedDict['tx']), self.lan5g['20rx'].append(speedDict['rx'])",
            'CH149 20M': "self.lan5g['20tx'].append(speedDict['tx']), self.lan5g['20rx'].append(speedDict['rx'])",
            'CH165 20M': "self.lan5g['20tx'].append(speedDict['tx']), self.lan5g['20rx'].append(speedDict['rx'])",
            'CH36 40M': "self.lan5g['40tx'].append(speedDict['tx']), self.lan5g['40rx'].append(speedDict['rx'])",
            'CH44 40M': "self.lan5g['40tx'].append(speedDict['tx']), self.lan5g['40rx'].append(speedDict['rx'])",
            'CH52 40M': "self.lan5g['40tx'].append(speedDict['tx']), self.lan5g['40rx'].append(speedDict['rx'])",
            'CH60 40M': "self.lan5g['40tx'].append(speedDict['tx']), self.lan5g['40rx'].append(speedDict['rx'])",
            'CH149 40M': "self.lan5g['40tx'].append(speedDict['tx']), self.lan5g['40rx'].append(speedDict['rx'])",
            'CH157 40M': "self.lan5g['40tx'].append(speedDict['tx']), self.lan5g['40rx'].append(speedDict['rx'])",
            'CH36 80M': "self.lan5g['80tx'].append(speedDict['tx']), self.lan5g['80rx'].append(speedDict['rx'])",
            'CH52 80M': "self.lan5g['80tx'].append(speedDict['tx']), self.lan5g['80rx'].append(speedDict['rx'])",
            'CH149 80M': "self.lan5g['80tx'].append(speedDict['tx']), self.lan5g['80rx'].append(speedDict['rx'])",
        }
        self.wan = {
            'CH11': "self.wan2g['tx'].append(speedDict['tx']), self.wan2g['rx'].append(speedDict['rx'])",
            'CH11 WRELAY': "self.wan2g['wrelaytx'].append(speedDict['tx']), self.wan2g['wrelayrx'].append(speedDict['rx'])",
            'CH11 RELAY': "self.wan2g['relaytx'].append(speedDict['tx']), self.wan2g['relayrx'].append(speedDict['rx'])",
            'GUEST': "self.wan2g['guesttx'].append(speedDict['tx']), self.wan2g['guestrx'].append(speedDict['rx'])",
            'CH149': "self.wan5g['tx'].append(speedDict['tx']), self.wan5g['rx'].append(speedDict['rx'])",
            'CH149 WRELAY': "self.wan5g['wrelaytx'].append(speedDict['tx']), self.wan5g['wrelayrx'].append(speedDict['rx'])",
            'CH149 RELAY': "self.wan5g['relaytx'].append(speedDict['tx']), self.wan5g['relayrx'].append(speedDict['rx'])",
        }

    def run(self):
        indexList = list() # iperf/ookla speedtest file
        report = open(self.reportName)
        for line in report:
            if not line.isspace():
                # m = re.search('AP_(.*)_CHAN(\d{1,3})_BW(\d{2})_((.*)_)?THROUGHPUT', line)
                mPattern = re.compile('AP_(.*)_CHAN(\d{1,3})_BW(\d{2})_((.*)_)?THROUGHPUT')
                nPattern = re.compile(r'AP_GUEST_([A-Z1-9]*)_OOKLA')
                oPattern = re.compile(r'AP_RELAY_([A-Z11-9]*)_CHAN(\d{1,3})_OOKLA')
                pPattern = re.compile(r'AP_WIRELESS_RELAY_([A-Z1-9]*)_CHAN(\d{1,3})_OOKLA')
                qPattern = re.compile(r'AP_([A-Z1-9]*)_CHAN(\d{1,3})_OOKLA')
                rPattern = re.compile(r'AP_WAN_BANDWIDTH')

                m = mPattern.search(line)
                if m:
                    logFile = self.logPath + m.group(0) + ".log"
                    chanBW = "CH" + m.group(2) + " " + m.group(3) + "M" # CH36 20M
                    encrypto = m.group(1) # CLEAR or PSK2
                    sheet = m.group(5) # infoTuple[3] is sheet name
                    infoTuple = (logFile, chanBW, encrypto, sheet)
                    if infoTuple not in indexList:
                        indexList.append(infoTuple)
                    continue

                n = nPattern.search(line)
                if n:
                    logFile = self.logPath + n.group(0) + ".log"
                    encrypto = n.group(1)
                    rf = 'GUEST'
                    module = ''
                    infoTuple = (logFile, rf + module, encrypto, 'WAN')
                    if infoTuple not in indexList:
                        indexList.append(infoTuple)
                    continue
                o = oPattern.search(line)
                if o:
                    logFile = self.logPath + o.group(0) + '.log'
                    encrypto = o.group(1)
                    rf = 'CH' + o.group(2)
                    module = ' RELAY'
                    infoTuple = (logFile, rf + module, encrypto, 'WAN')
                    if infoTuple not in indexList:
                        indexList.append(infoTuple)
                    continue

                p = pPattern.search(line)
                if p:
                    logFile = self.logPath + p.group(0) + '.log'
                    encrypto = p.group(1)
                    rf = 'CH' + p.group(2)
                    module = ' WRELAY'
                    infoTuple = (logFile, rf + module, encrypto, 'WAN')
                    if infoTuple not in indexList:
                        indexList.append(infoTuple)
                    continue

                q = qPattern.search(line)
                if q:
                    logFile = self.logPath + q.group(0) + '.log'
                    encrypto = q.group(1)
                    rf = 'CH' + q.group(2)
                    module = ''
                    infoTuple = (logFile, rf + module, encrypto, 'WAN')
                    if infoTuple not in indexList:
                        indexList.append(infoTuple)
                    continue

                r = rPattern.search(line)
                if r:
                    logFile = self.logPath + r.group(0) + '.log'
                    wanBW = GetWanBandwidth(logFile)
                    wanBW.start()
                    wanBW.join()
                    self.wanBWResult = wanBW.result
                    continue

        report.close()
        if len(indexList) is not 0:
            try:
                wb = load_workbook(MAIL_THROUGHPUT_XLSX_ORIGINAL)
            except:
                print "specified file no exists!"
                return
            for tu in indexList:
                # collect draw chart data
                if tu[3] == "DUT":
                    speedDict = self.getIperfThroughputLogVerbose(tu[0])
                    eval(self.dut.get(tu[1]))
                elif tu[3] == "LAN":
                    speedDict = self.getIperfThroughputLogVerbose(tu[0])
                    eval(self.lan.get(tu[1]))
                elif tu[3] == "WAN":
                    speedDict = self.getOoklaThroughputLogVerbose(tu[0])
                    eval(self.wan.get(tu[1]))

                for cell in wb[tu[3]].get_cell_collection():
                    if tu[1] == cell.value:
                        x = cell.row
                        y = column_index_from_string(cell.column)
                        if tu[2] == "PSK2":
                            x += 3
                            y += 1
                            while wb[tu[3]].cell(row=x, column=y).value is not None:
                                x += 1
                            wb[tu[3]].cell(row=x, column=y).value = speedDict["tx"]
                        elif tu[2] == "CLEAR":
                            x += 3
                            y += 2
                            while wb[tu[3]].cell(row=x, column=y).value is not None:
                                x += 1
                            wb[tu[3]].cell(row=x, column=y).value = speedDict["tx"]
                        y += 2 # downlink cell.column + 2 = uplink cell.column
                        wb[tu[3]].cell(row=x, column=y).value = speedDict['rx']
                        break
            wb.save(MAIL_THROUGHPUT_XLSX)

            # for key, value in self.dut2g.iteritems():
            #     i = 0
            #     count = len(value)
            #     while i < count:
            #         if not isfloat(value[i]):
            #             value.pop(i)
            #             count = len(value)
            #         else:
            #             i += 1
            #     if len(value) is not 0:
            #         ave = round(float(reduce(lambda i, j: float(i)+float(j), value))/len(value), 2)
            #         self.dut2g[key] = ave
            #     else:
            #         self.dut2g[key] = 0

            self.dut2g = self.getAveSpeed(self.dut2g)
            self.dut5g = self.getAveSpeed(self.dut5g)
            self.lan2g = self.getAveSpeed(self.lan2g)
            self.lan5g = self.getAveSpeed(self.lan5g)
            self.wan2g = self.getAveSpeed(self.wan2g)
            self.wan5g = self.getAveSpeed(self.wan5g)

            # draw chart
            self.drawThroughput2g(self.dut2g, MAIL_PIC2)
            self.drawThroughput5g(self.dut5g, MAIL_PIC3)
            self.drawThroughput2g(self.lan2g, MAIL_PIC5)
            self.drawThroughput5g(self.lan5g, MAIL_PIC6)
            self.drawOoklaThroughput2g(self.wan2g, MAIL_PIC7)
            self.drawOoklaThroughput5g(self.wan5g, MAIL_PIC8)


    def getAveSpeed(self, speeddict):
        for key, value in speeddict.iteritems():
            i = 0
            count = len(value)
            while i < count:
                if not self.isfloat(value[i]):
                    value.pop(i)
                    count = len(value)
                else:
                    i += 1
            if len(value) is not 0:
                ave = round(float(reduce(lambda i, j: float(i)+float(j), value))/len(value), 2)
                speeddict[key] = ave
            else:
                speeddict[key] = 0
        return speeddict

    def isfloat(self, value):
      try:
        float(value)
        return True
      except ValueError:
        return False

    def getOoklaThroughputLogVerbose(self, logfile):
        result = {
            'tx': 0,
            'rx': 0,
        }
        try:
            log = open(logfile)
            for line in log:
                if not line.isspace():
                    mPattern = re.compile('downlink rate: (.*) Mbps, uplink rate: (.*) Mbps')
                    m = mPattern.search(line)
                    if m:
                        result['tx'] = m.group(1)
                        result['rx'] = m.group(2)
                        break
            log.close()
        except IOError as e:
            raise  e
        return result


    def getIperfThroughputLogVerbose(self, logfile):
        result = {
            'tx': 0,
            'rx': 0,
        }
        try:
            log = open(logfile)
            for line in log:
                if not line.isspace():
                    m = re.search('0\.0-\d{1,4}.*\s(\d{1,}\.?\d{1,})?\sMbits/sec', line)
                    if m:
                        if result['tx'] is 0:
                            result['tx'] = m.group(1)
                            continue
                        else:
                            result['rx'] = m.group(1)
                            break
            log.close()
        except IOError as e:
            raise e
        if float(result['tx']) >= WIFI_MAX_THROUGHPUT:
            result['tx'] = 'BREAK'
        elif float(result['tx']) == 0:
            result['tx'] = 'ERROR'
        if float(result['rx']) >= WIFI_MAX_THROUGHPUT:
            result['rx'] = 'BREAK'
        elif float(result['rx']) == 0:
            result['rx'] = 'ERROR'

        return result


    def drawThroughput2g(self, data, picname):
        ret = data
        for value in ret.values():
            if isinstance(value, float):
                bar_width = 0.4
                opacity = 0.4
                index = np.arange(2)
                tx = list()
                rx = list()
                tx.append(ret.get("20tx"))
                tx.append(ret.get("40tx"))
                rx.append(ret.get("20rx"))
                rx.append(ret.get("40rx"))

                fig, ax = plt.subplots(figsize=(6, 4))
                print "draw %s chart" % picname
                # plt.subplots_adjust(left=0.08, right=0.95)
                rects1 = plt.bar(index, tx, bar_width,
                                 alpha=opacity,
                                 color='b',
                                 label='Tx',
                                 )

                rects2 = plt.bar(index + bar_width, rx, bar_width,
                                 alpha=opacity,
                                 color='r',
                                 label='Rx',
                                 )

                def autolabel(rects):
                    # attach some text labels
                    for rect in rects:
                        height = rect.get_height()
                        ax.text(rect.get_x()+rect.get_width()/2., 0.75*height, '%.1f'%float(height),
                                ha='center', va='bottom', fontsize=10)

                autolabel(rects1)
                autolabel(rects2)

                # plt.xlabel('Bandwidth', fontsize=10)
                plt.ylabel('Mbps', fontsize=10)
                plt.suptitle(picname.split(".")[0].split('\\')[-1].title(), fontsize=10)
                plt.xticks(index + bar_width, ('20MHz', '40MHz',), fontsize=10)
                plt.yticks(fontsize=10)
                plt.legend(prop={'size': 10})

                # plt.show()
                plt.savefig(picname)
                plt.close()
                break


    def drawThroughput5g(self, data, picname):
        ret = data
        for value in ret.values():
            if isinstance(value, float):
                bar_width = 0.4
                opacity = 0.4
                index = np.arange(3)
                tx = list()
                rx = list()
                tx.append(ret.get("20tx"))
                tx.append(ret.get("40tx"))
                tx.append(ret.get("80tx"))
                rx.append(ret.get("20rx"))
                rx.append(ret.get("40rx"))
                rx.append(ret.get("80rx"))

                fig, ax = plt.subplots(figsize=(6, 4))
                print "draw %s chart" % picname
                # plt.subplots_adjust(left=0.08, right=0.95)
                rects1 = plt.bar(index, tx, bar_width,
                                 alpha=opacity,
                                 color='b',
                                 label='Tx'
                                 )

                rects2 = plt.bar(index + bar_width, rx, bar_width,
                                 alpha=opacity,
                                 color='r',
                                 label='Rx'
                                 )

                def autolabel(rects):
                    # attach some text labels
                    for rect in rects:
                        height = rect.get_height()
                        ax.text(rect.get_x()+rect.get_width()/2., 0.75*height, '%.1f'%float(height),
                                ha='center', va='bottom', fontsize=10)

                autolabel(rects1)
                autolabel(rects2)

                # plt.xlabel('Bandwidth', fontsize=10)
                plt.ylabel('Mbps', fontsize=10)
                plt.suptitle(picname.split(".")[0].split('\\')[-1].title(), fontsize=10)
                plt.xticks(index + bar_width, ('20MHz', '40MHz', '80MHz'), fontsize=10)
                plt.yticks(fontsize=10)
                plt.legend(prop={'size': 10})

                # plt.show()
                plt.savefig(picname)
                plt.close()
                break


    def drawOoklaThroughput2g(self, data, picname):
        ret = data
        for value in ret.values():
            if isinstance(value, float):
                bar_width = 0.4
                opacity = 0.4
                index = np.arange(4)
                tx = list()
                rx = list()
                tx.append(ret.get("tx"))
                tx.append(ret.get("guesttx"))
                tx.append(ret.get("wrelaytx"))
                tx.append(ret.get("relaytx"))

                rx.append(ret.get("rx"))
                rx.append(ret.get("guestrx"))
                rx.append(ret.get("wrelayrx"))
                rx.append(ret.get("relayrx"))
                fig, ax = plt.subplots(figsize=(6, 4))
                print "draw %s chart" % picname
                # plt.subplots_adjust(left=0.08, right=0.95)
                rects1 = plt.bar(index, tx, bar_width,
                                 alpha=opacity,
                                 color='b',
                                 label='Tx'
                                 )

                rects2 = plt.bar(index + bar_width, rx, bar_width,
                                 alpha=opacity,
                                 color='r',
                                 label='Rx'
                                 )

                def autolabel(rects):
                    # attach some text labels
                    for rect in rects:
                        height = rect.get_height()
                        ax.text(rect.get_x()+rect.get_width()/2., 0.75*height, '%.1f'%float(height),
                                ha='center', va='bottom', fontsize=10)

                autolabel(rects1)
                autolabel(rects2)

                # plt.xlabel('Bandwidth', fontsize=10)
                plt.ylabel('Mbps', fontsize=10)
                plt.suptitle(picname.split(".")[0].split('\\')[-1].title(), fontsize=10)
                plt.xticks(index + bar_width, ('Normal', 'Guest', 'Wireless-Relay', 'Relay'), fontsize=10)
                plt.yticks(fontsize=10)
                plt.legend(prop={'size': 10})

                # plt.show()
                plt.savefig(picname)
                plt.close()
                break


    def drawOoklaThroughput5g(self, data, picname):
        ret = data
        for value in ret.values():
            if isinstance(value, float):
                bar_width = 0.4
                opacity = 0.4
                index = np.arange(3)
                tx = list()
                rx = list()
                tx.append(ret.get("tx"))
                tx.append(ret.get("wrelaytx"))
                tx.append(ret.get("relaytx"))

                rx.append(ret.get("rx"))
                rx.append(ret.get("wrelayrx"))
                rx.append(ret.get("relayrx"))
                fig, ax = plt.subplots(figsize=(6, 4))
                print "draw %s chart" % picname
                # plt.subplots_adjust(left=0.08, right=0.95)
                rects1 = plt.bar(index, tx, bar_width,
                                 alpha=opacity,
                                 color='b',
                                 label='Tx'
                                 )

                rects2 = plt.bar(index + bar_width, rx, bar_width,
                                 alpha=opacity,
                                 color='r',
                                 label='Rx'
                                 )

                def autolabel(rects):
                    # attach some text labels
                    for rect in rects:
                        height = rect.get_height()
                        ax.text(rect.get_x()+rect.get_width()/2., 0.75*height, '%.1f'%float(height),
                                ha='center', va='bottom', fontsize=10)

                autolabel(rects1)
                autolabel(rects2)

                # plt.xlabel('Bandwidth', fontsize=10)
                plt.ylabel('Mbps', fontsize=10)
                plt.suptitle(picname.split(".")[0].split('\\')[-1].title(), fontsize=10)
                plt.xticks(index + bar_width, ('Normal', 'Wireless-Relay', 'Relay'), fontsize=10)
                plt.yticks(fontsize=10)
                plt.legend(prop={'size': 10})

                # plt.show()
                plt.savefig(picname)
                plt.close()
                break


class GetOnlineLog(threading.Thread):
    def __init__(self, report):
        threading.Thread.__init__(self)
        self.running = False
        self.reportName = report
        self.logPath = TEST_SUITE_LOG_PATH
        self.result = {
            "pass": 0,
            "fail": 0,
            }

    def run(self):
        self.running = True
        logFileList = list()
        report = open(self.reportName)
        for line in report:
            if not line.isspace():
                m = re.search('assoc_repeat.*testcase.*\.(.*)\)', line)
                if m:
                    if line.endswith("ok\n"):
                        self.result["pass"] += 100
                    elif line.endswith("FAIL\n"):
                        logFile = m.group(1) + ".log"
                        logFile = self.logPath + logFile
                        if logFile not in logFileList:
                            logFileList.append(logFile)
                else:
                    n = re.search('assoc', line)
                    if n:
                        if line.endswith("ok\n"):
                            self.result["pass"] += 1
                        elif line.endswith("FAIL\n"):
                            self.result["fail"] += 1
        report.close()
        for lf in logFileList:
            log = open(lf)
            for line in log:
                if not line.isspace():
                    m = re.search('Not all association were successful.*but was:<(\d+)>', line)
                    if m:
                        self.result["pass"] += int(m.group(1))
                        self.result["fail"] += (100 - int(m.group(1)))
            log.close()
        self.stop()

    def stop(self):
        self.running = False


class GetTestModule(threading.Thread):
    def __init__(self, report):
        threading.Thread.__init__(self)
        self.running = False
        self.reportName = report
        self.logPath = TEST_SUITE_LOG_PATH
        self.moduleDict = {
            'treeBasicApi': '基础功能',
            'treeGuestWifiApi': '访客网络',
            'treeBSDApi': '双频合一',
            'treeWireRelayApi': '有线中继',
            'treeWirelessRelayApi': '无线中继',
            'treeAccessControlApi': '接入控制',
            'treeQosApi': '智能限速',
            'treeThroughputDUTApi': 'DUT2WiFi吞吐',
            'treeThroughputLANApi': 'LAN2WiFi吞吐',
            'treeThroughputWANApi': 'WAN2WiFi吞吐',
            'treeStressApi': '压力测试',
            'treeOthersApi': '其他',
        }
        self.result = list()

    def run(self):
        self.running = True
        f = open(self.reportName)
        for line in f:
            if not line.isspace():
                m = re.search('\(testcase.*\.(.*)\)', line)
                if m:
                    testCase = m.group(1)
                    for module in self.moduleDict.iterkeys():
                        if self.chkTestCaseModule(testCase, module):
                            if self.moduleDict.get(module) not in self.result:
                                self.result.append(self.moduleDict.get(module))
                                self.result.append('，')
        if len(self.result) > 0:
            del self.result[-1]
        f.close()
        self.stop()

    def stop(self):
        self.running = False


    def chkTestCaseModule(self, tcName, module):
        for i in getattr(data, module):
            if isinstance(i, str):
                if tcName == i:
                    return True
        return False


class GetWanBandwidth(threading.Thread):
    def __init__(self, report):
        threading.Thread.__init__(self)
        self.running = False
        self.reportName = report
        self.logPath = TEST_SUITE_LOG_PATH
        self.result = {
            "wandownload": 0,
            "wanupload": 0,
        }

    def run(self):
        self.running = True
        f = open(self.reportName)
        mPattern = re.compile('\"bandwidth2\":(\d{1,3}\.\d{1,3})') # bandwidth2 means upload
        nPattern = re.compile('\"bandwidth\":(\d{1,3}\.\d{1,3})') # bandwidth means download
        for line in f:
            if not line.isspace():
                m = mPattern.search(line)
                n = nPattern.search(line)
                if m and n:
                    self.result['wandownload'] = n.group(1)
                    self.result['wanupload'] = m.group(1)
                    break


if __name__ == '__main__':
    # print getThroughputLogVerbose("D:\Python\peanuts\AP_CLEAR_CHAN36_BW20_LAN_THROUGHPUT.log")
    # print getChannelFlowLogVerbose("E:\peanuts\AP_MIXEDPSK_CHAN1_36_FLOW.log")
    info = GetThroughputLog("R1D 稳定版 2.14.5.log".decode("utf8").encode("gbk"))
    info.start()
    # info.join()
    # t = GetTestModule("R1CM 开发版 2.11.13.log".decode('utf8').encode('gbk'))
    # t.start()





