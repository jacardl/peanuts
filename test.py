# import pygal                                                       # First import pygal
# bar_chart = pygal.Line()                                            # Then create a bar graph object
# bar_chart.add('Fibonacci', [0, 1, 1, 2, 3, 5, 8, 13, 21, 34, 55])  # Add some values
# bar_chart.render_to_png('bar_chart.png')

# from urllib import urlopen
# data = urlopen('http://peak.telecommunity.com/dist/ez_setup.py')
# with open('ez_setup.py', 'wb') as f:
#     f.write(data.read())

import os
# file_name = 'ez_setup.py'
# from urllib import urlopen
# data = urlopen('http://peak.telecommunity.com/dist/ez_setup.py')
# with open(file_name, 'wb') as f:
#     f.write(data.read())
# os.system("D:\\Python\\\new peanuts\\24\\ez_setup.py")
# import numpy as np
# import matplotlib.pyplot as plt
#
#
# bar_width = 0.42
# opacity = 0.4
# index = np.arange(6)
# ret = {'tx2gClear': 50.9, 'rx2gTkip': 17.7, 'tx5gAes': 93.27, 'rx2gClear': 42.4, 'tx5gTkip': 26.4, 'rx2gAes': 39.1,
#        'tx5gClear': 93.6, 'tx2gAes': 47.83, 'rx5gTkip': 19.8, 'rx5gAes': 52.2, 'tx2gTkip': 26.45, 'rx5gClear': 57.4}
#
# tx = list()
# rx = list()
# tx.append(ret.get("tx2gClear"))
# tx.append(ret.get("tx2gAes"))
# tx.append(ret.get("tx2gTkip"))
# tx.append(ret.get("tx5gClear"))
# tx.append(ret.get("tx5gAes"))
# tx.append(ret.get("tx5gTkip"))
# rx.append(ret.get("rx2gClear"))
# rx.append(ret.get("rx2gAes"))
# rx.append(ret.get("rx2gTkip"))
# rx.append(ret.get("rx5gClear"))
# rx.append(ret.get("rx5gAes"))
# rx.append(ret.get("rx5gTkip"))
#
#
# fig, ax = plt.subplots()
# rects1 = plt.bar(index, tx, bar_width,
#                  alpha=opacity,
#                  color='b',
#                  label='Tx'
#                  )
#
# rects2 = plt.bar(index + bar_width, rx, bar_width,
#                  alpha=opacity,
#                  color='r',
#                  label='Rx'
#                  )
#
# def autolabel(rects):
#     # attach some text labels
#     for rect in rects:
#         height = rect.get_height()
#         ax.text(rect.get_x()+rect.get_width()/2., 0.75*height, '%.1f'%float(height),
#                 ha='center', va='bottom')
#
# autolabel(rects1)
# autolabel(rects2)
#
#
# plt.xlabel('Radio & cipher suite')
# plt.ylabel('Mbps')
# plt.title('Throughput')
# plt.xticks(index + bar_width, ('2.4g_Clear', '2.4g_AES', '2.4g_TKIP', '5g_Clear', '5g_AES', '5g_TKIP',))
# plt.legend()
#
# # plt.show()
# plt.savefig("Throughput")
# plt.close()
from openpyxl import Workbook
from openpyxl import load_workbook

fileName = "MEM_2015.10.22 20.55.16.xlsx"
wb = load_workbook(fileName)
ws = wb.active
maxRow = ws.get_highest_row()
maxCol = ws.get_highest_column()

for col in range(2, maxCol + 1):
    rowStart = None
    rowEnd = None
    for row1 in range(2, maxRow+1):
        if rowStart is None:
            rowStart = ws.cell(row=row1, column=col).value
        else:
            rowStart = float(rowStart)
            break
    for row2 in reversed(range(2, maxRow+1)):
        if rowEnd is None:
            rowEnd = ws.cell(row=row2, column=col).value
        else:
            rowEnd = float(rowEnd)
            break
    try:
        diff = (rowEnd - rowStart) / rowStart
    except Exception, e:
        print e, "rowStart=%s, rowEnd=%s"%rowStart,rowEnd

    if diff != 0:
        # self.ws.write(maxRow, col, diff, stylePercent)
        sum = ws.cell(row=maxRow + 1, column=col)
        diff = "{:.1%}".format(diff)
        sum.value = diff
        # else:
        # # self.ws.write(maxRow, col, diff)
        # sum = self.ws.cell(row=maxRow+1, column=col)
        # sum.value = diff

wb.save(fileName)
